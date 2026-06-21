from __future__ import annotations

from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

TARGET = ROOT / "transitions_teammate_complete_graph_v4_executable.xlsx"
CHAIN_CSV = ROOT / "verification_report_static_bridge_chain.csv"
REPORT = ROOT / "verification_report_patch_teammate_state_after_safe.md"

DIRECTION_MARKERS = [
    "лицом к чужим воротам",
    "лицом к своим воротам",
    "лицом к боковой линии",
]


def clean(value) -> str:
    return str(value or "").strip()


def norm(value) -> str:
    return clean(value).lower().replace("ё", "е")


def split_position_direction(value: str) -> tuple[str, str]:
    text = clean(value)
    for marker in DIRECTION_MARKERS:
        tail = ", " + marker
        if text.endswith(tail):
            return text[: -len(tail)].strip(), marker
    return text, ""


def headers(ws) -> dict[str, int]:
    return {
        clean(ws.cell(row=1, column=col).value): col
        for col in range(1, ws.max_column + 1)
        if clean(ws.cell(row=1, column=col).value)
    }


def ensure_column(ws, h: dict[str, int], name: str) -> int:
    if name in h:
        return h[name]
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = name
    h[name] = col
    return col


def get_cell(ws, row: int, h: dict[str, int], name: str) -> str:
    col = h.get(name)
    if not col:
        return ""
    return clean(ws.cell(row=row, column=col).value)


def set_cell(ws, row: int, h: dict[str, int], name: str, value: str, changes: Counter) -> None:
    col = h[name]
    value = clean(value)
    old = clean(ws.cell(row=row, column=col).value)
    if old != value:
        ws.cell(row=row, column=col).value = value
        changes[name] += 1


def main() -> int:
    if not TARGET.exists():
        raise FileNotFoundError(f"Target file not found: {TARGET}")
    if not CHAIN_CSV.exists():
        raise FileNotFoundError(
            f"Required audit CSV not found: {CHAIN_CSV.name}. "
            "Run: python tools\\audit_static_bridge_chain.py"
        )

    audit = pd.read_csv(CHAIN_CSV, dtype=str, keep_default_na=False)
    required_cols = {
        "excel_row",
        "status",
        "ball_holder_position_consensus",
        "ball_holder_direction_consensus",
        "transition_id",
    }
    missing_cols = sorted(required_cols - set(audit.columns))
    if missing_cols:
        raise ValueError(f"Audit CSV missing columns: {missing_cols}")

    safe_by_excel_row = {}
    for _, row in audit.iterrows():
        if clean(row.get("status")) != "PATCH_SAFE":
            continue
        excel_row = int(row["excel_row"])
        bh_pos = clean(row.get("ball_holder_position_consensus"))
        bh_dir = clean(row.get("ball_holder_direction_consensus"))
        if not bh_pos or not bh_dir:
            raise ValueError(f"PATCH_SAFE row without ball_holder consensus: Excel row {excel_row}")
        safe_by_excel_row[excel_row] = {
            "transition_id": clean(row.get("transition_id")),
            "ball_holder_position_after": bh_pos,
            "ball_holder_direction_after": bh_dir,
        }

    wb = load_workbook(TARGET)
    ws = wb.active
    h = headers(ws)

    original_column_count = ws.max_column
    original_row_count = ws.max_row - 1

    for name in [
        "player_direction_after",
        "ball_holder_position_after",
        "ball_holder_direction_after",
    ]:
        ensure_column(ws, h, name)

    changes = Counter()
    counts = Counter()
    attention_examples = []

    for excel_row in range(2, ws.max_row + 1):
        transition_id = get_cell(ws, excel_row, h, "transition_id")

        raw_player_position = get_cell(ws, excel_row, h, "player_position_after")
        parsed_position, parsed_direction = split_position_direction(raw_player_position)
        if parsed_direction:
            existing_direction = get_cell(ws, excel_row, h, "player_direction_after")
            if existing_direction and norm(existing_direction) != norm(parsed_direction):
                counts["skip_player_direction_conflict"] += 1
                if len(attention_examples) < 80:
                    attention_examples.append(
                        (excel_row, transition_id, "PLAYER_DIRECTION_CONFLICT", existing_direction, parsed_direction)
                    )
            else:
                set_cell(ws, excel_row, h, "player_position_after", parsed_position, changes)
                set_cell(ws, excel_row, h, "player_direction_after", parsed_direction, changes)
                counts["patched_player_position_direction"] += 1
        else:
            counts["no_embedded_player_direction"] += 1

        safe = safe_by_excel_row.get(excel_row)
        if not safe:
            counts["skip_ball_holder_not_patch_safe"] += 1
            continue

        expected_transition_id = safe["transition_id"]
        if expected_transition_id and transition_id != expected_transition_id:
            counts["skip_ball_holder_transition_id_mismatch"] += 1
            if len(attention_examples) < 80:
                attention_examples.append(
                    (excel_row, transition_id, "TRANSITION_ID_MISMATCH", transition_id, expected_transition_id)
                )
            continue

        new_bh_pos = safe["ball_holder_position_after"]
        new_bh_dir = safe["ball_holder_direction_after"]
        existing_bh_pos = get_cell(ws, excel_row, h, "ball_holder_position_after")
        existing_bh_dir = get_cell(ws, excel_row, h, "ball_holder_direction_after")

        if existing_bh_pos and norm(existing_bh_pos) != norm(new_bh_pos):
            counts["skip_ball_holder_position_conflict"] += 1
            if len(attention_examples) < 80:
                attention_examples.append(
                    (excel_row, transition_id, "BALL_HOLDER_POSITION_CONFLICT", existing_bh_pos, new_bh_pos)
                )
            continue

        if existing_bh_dir and norm(existing_bh_dir) != norm(new_bh_dir):
            counts["skip_ball_holder_direction_conflict"] += 1
            if len(attention_examples) < 80:
                attention_examples.append(
                    (excel_row, transition_id, "BALL_HOLDER_DIRECTION_CONFLICT", existing_bh_dir, new_bh_dir)
                )
            continue

        set_cell(ws, excel_row, h, "ball_holder_position_after", new_bh_pos, changes)
        set_cell(ws, excel_row, h, "ball_holder_direction_after", new_bh_dir, changes)
        counts["patched_ball_holder_patch_safe"] += 1

    final_column_count = ws.max_column
    wb.save(TARGET)
    wb.close()

    lines = [
        "# Safe Patch Teammate State-After Report",
        "",
        "Status: PATCH_APPLIED_CONSERVATIVE",
        "",
        f"- target_file: {TARGET.name}",
        f"- audit_csv: {CHAIN_CSV.name}",
        f"- rows_scanned: {original_row_count}",
        f"- patch_safe_rows_from_audit: {len(safe_by_excel_row)}",
        f"- original_column_count: {original_column_count}",
        f"- final_column_count: {final_column_count}",
        "",
        "## Changed Cells by Column",
        "",
    ]
    for key, value in changes.most_common():
        lines.append(f"- {key}: {value}")

    lines.extend(["", "## Counts", ""])
    for key, value in counts.most_common():
        lines.append(f"- {key}: {value}")

    lines.extend(["", "## Attention Examples", ""])
    for excel_row, transition_id, flag, current, expected in attention_examples:
        lines.append(f"- row {excel_row} {transition_id}")
        lines.append(f"  flag: {flag}")
        lines.append(f"  current: {current}")
        lines.append(f"  expected: {expected}")

    REPORT.write_text("\\n".join(lines), encoding="utf-8")

    print("Safe teammate state-after patch status: PASS")
    print(f"Patched file: {TARGET.name}")
    print(f"Report written: {REPORT.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
