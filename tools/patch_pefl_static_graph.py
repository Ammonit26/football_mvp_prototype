from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

SCENE_LIBRARIES = {
    "PLAYER_WITH_BALL": (ROOT / "ball_at_player_normalized_prototype_v7_executable.xlsx", "normalized_player_scenes"),
    "TEAMMATE_WITH_BALL": (ROOT / "ball_at_teammate_normalized_prototype_v1_executable.xlsx", "normalized_teammate_scenes"),
    "OPPONENT_WITH_BALL": (ROOT / "ball_at_opponent_normalized_prototype_v2_executable.xlsx", "normalized_opponent_scenes"),
}

TRANSITION_LIBRARIES = {
    "player": ROOT / "transitions_player_complete_graph_v2_executable.xlsx",
    "teammate": ROOT / "transitions_teammate_complete_graph_v4_executable.xlsx",
    "opponent": ROOT / "transitions_opponent_complete_graph_v1_executable.xlsx",
}

CONTINUE_ACTION = "CONTINUE"
CONTINUE_OUTCOME = "SUCCESS"
STATIC_PREFIX = "fb_static_auto_"
STATIC_TRANSITION_PREFIX = "fb_trans_static_auto_"


@dataclass(frozen=True)
class SceneRecord:
    scene_id: str
    owner: str
    player_position: str
    player_direction: str
    ball_holder_position: str
    ball_holder_direction: str
    ball_holder_action: str
    available_player_actions: str
    narrative_scene: str
    scene_type: str


@dataclass(frozen=True)
class TransitionPatch:
    library_key: str
    source_row: int
    source_transition_id: str
    source_scene_id: str
    source_owner: str
    action: str
    outcome: str
    ball_owner_after: str
    player_position_after: str
    player_direction_after: str
    ball_holder_position_after: str
    ball_holder_direction_after: str
    original_next_scene_id: str
    original_allowed_next_scene_ids: str
    static_scene_id: str
    static_transition_id: str
    static_text: str


def safe(value: object) -> str:
    return "" if value is None else str(value).strip()


def split_pool(raw: object) -> List[str]:
    return [part.strip() for part in safe(raw).split("||") if part.strip()]


def join_pool(values: Iterable[str]) -> str:
    out: List[str] = []
    for value in values:
        if value and value not in out:
            out.append(value)
    return "||".join(out)


def stable_suffix(*parts: str, length: int = 12) -> str:
    raw = "|".join(parts).encode("utf-8")
    return hashlib.sha1(raw).hexdigest()[:length]


def header_map(ws) -> Dict[str, int]:
    return {safe(cell.value): int(cell.column) for cell in ws[1] if safe(cell.value)}


def require(headers: Dict[str, int], names: Iterable[str]) -> int:
    lowered = {name.lower(): col for name, col in headers.items()}
    for name in names:
        if name.lower() in lowered:
            return lowered[name.lower()]
    raise RuntimeError(f"Missing required column. Tried: {list(names)}")


def optional(headers: Dict[str, int], names: Iterable[str]) -> Optional[int]:
    lowered = {name.lower(): col for name, col in headers.items()}
    for name in names:
        if name.lower() in lowered:
            return lowered[name.lower()]
    return None


def ensure_column(ws, name: str) -> int:
    headers = header_map(ws)
    if name in headers:
        return headers[name]
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = name
    return col


def find_rows(ws, col: int, value: str) -> list[int]:
    return [row for row in range(2, ws.max_row + 1) if safe(ws.cell(row=row, column=col).value) == value]


def set_if_present(ws, headers: Dict[str, int], row: int, column_name: str, value: object) -> None:
    col = optional(headers, [column_name])
    if col is not None:
        ws.cell(row=row, column=col).value = value


def load_scene_records() -> Dict[str, SceneRecord]:
    scenes: Dict[str, SceneRecord] = {}
    for owner, (path, sheet_name) in SCENE_LIBRARIES.items():
        wb = load_workbook(path)
        ws = wb[sheet_name]
        ensure_column(ws, "scene_type")
        headers = header_map(ws)
        scene_id_col = require(headers, ["scene_id"])
        scene_type_col = require(headers, ["scene_type"])
        changed = False
        for row in range(2, ws.max_row + 1):
            if not safe(ws.cell(row=row, column=scene_type_col).value):
                ws.cell(row=row, column=scene_type_col).value = "dynamic"
                changed = True
            scene_id = safe(ws.cell(row=row, column=scene_id_col).value)
            if not scene_id:
                continue
            scenes[scene_id] = SceneRecord(
                scene_id=scene_id,
                owner=safe(ws.cell(row=row, column=optional(headers, ["scene_owner"]) or 0).value) or owner,
                player_position=safe(ws.cell(row=row, column=optional(headers, ["player_position"]) or 0).value),
                player_direction=safe(ws.cell(row=row, column=optional(headers, ["player_direction"]) or 0).value),
                ball_holder_position=safe(ws.cell(row=row, column=optional(headers, ["ball_holder_position"]) or 0).value),
                ball_holder_direction=safe(ws.cell(row=row, column=optional(headers, ["ball_holder_direction"]) or 0).value),
                ball_holder_action=safe(ws.cell(row=row, column=optional(headers, ["ball_holder_action"]) or 0).value),
                available_player_actions=safe(ws.cell(row=row, column=optional(headers, ["available_player_actions"]) or 0).value),
                narrative_scene=safe(ws.cell(row=row, column=optional(headers, ["narrative_scene"]) or 0).value),
                scene_type=safe(ws.cell(row=row, column=scene_type_col).value).lower() or "dynamic",
            )
        if changed:
            wb.save(path)
            print(f"DEFAULT_SCENE_TYPE_FILLED {path.name}")
    return scenes


def action_bucket(action: str) -> str:
    lower = action.lower()
    if "удар" in lower or "бить" in lower or "бьет" in lower:
        return "shot"
    if "дриблинг" in lower or "пройти" in lower or "обыграть" in lower or "вести" in lower:
        return "carry_dribble"
    if "пас" in lower or "передач" in lower or "заброс" in lower or "навес" in lower or "прострел" in lower or "отдать" in lower or "сыграть" in lower:
        return "pass_cross"
    if "отбор" in lower or "накрыть" in lower or "перехват" in lower or "закрывать" in lower or "оттянуться" in lower or "зоне" in lower or "рывок" in lower:
        return "defensive_action"
    if "отскок" in lower or "рикошет" in lower or "подбор" in lower:
        return "second_ball"
    return "generic"


def static_text(source_scene: SceneRecord, action: str, outcome: str, owner_after: str) -> str:
    bucket = action_bucket(action)
    ok = outcome.upper() == "SUCCESS"
    if bucket == "shot":
        return "Удар завершает развитие атаки. Мяч меняет фазу эпизода, и игрок получает следующую ситуацию уже после реакции защиты и вратаря."
    if bucket == "carry_dribble" and ok:
        return "Действие проходит. Игрок продвигает мяч и заставляет соперника перестраиваться, прежде чем возникает следующая игровая ситуация."
    if bucket == "carry_dribble" and not ok:
        return "Соперник срывает продвижение. Мяч меняет владельца, и эпизод сразу переходит в новую фазу под давлением."
    if bucket == "pass_cross" and ok:
        return "Передача проходит. Мяч доходит до адресата, и атака получает продолжение уже в новой расстановке игроков."
    if bucket == "pass_cross" and not ok:
        return "Передача не проходит. Соперник читает направление мяча, и владение переходит к другой команде."
    if bucket == "defensive_action" and ok:
        return "Оборонительное действие срабатывает. Опасность частично снята, и следующая фаза начинается уже после этого вмешательства."
    if bucket == "defensive_action" and not ok:
        return "Попытка оборонительного действия не останавливает атаку. Соперник сохраняет темп и получает следующую ситуацию."
    if bucket == "second_ball" and ok:
        return "Игрок правильно читает отскок и успевает к мячу. Эпизод продолжается после борьбы за второй мяч."
    if bucket == "second_ball" and not ok:
        return "Борьба за отскок проиграна. Мяч достается сопернику, и направление эпизода меняется."
    if owner_after != source_scene.owner:
        return "Событие меняет владение мячом. Игрок видит новую фазу эпизода уже после смены инициативы."
    return "Событие меняет расположение игроков на поле. Следующая сцена начинается уже после этой промежуточной фазы."


def is_transition_already_wrapped(allowed_pool: List[str], scenes: Dict[str, SceneRecord]) -> bool:
    if not allowed_pool:
        return False
    return all(scene_id in scenes and scenes[scene_id].scene_type == "static" for scene_id in allowed_pool)


def collect_transition_patches(scenes: Dict[str, SceneRecord]) -> List[TransitionPatch]:
    patches: List[TransitionPatch] = []
    for library_key, path in TRANSITION_LIBRARIES.items():
        wb = load_workbook(path)
        ws = wb.active
        headers = header_map(ws)
        transition_id_col = optional(headers, ["transition_id"])
        source_col = require(headers, ["source_scene_id", "source_scene", "from_scene", "scene_id"])
        action_col = require(headers, ["player_action", "action", "teammate_action"])
        outcome_col = require(headers, ["player_action_outcome", "teammate_action_outcome", "outcome"])
        owner_col = require(headers, ["ball_owner_after", "ball_owner"])
        next_col = optional(headers, ["next_scene_id", "target_scene_id"])
        allowed_col = require(headers, ["allowed_next_scene_ids"])
        player_pos_after_col = optional(headers, ["player_position_after"])
        player_dir_after_col = optional(headers, ["player_direction_after"])
        ball_holder_pos_after_col = optional(headers, ["ball_holder_position_after"])
        ball_holder_dir_after_col = optional(headers, ["ball_holder_direction_after"])

        for row in range(2, ws.max_row + 1):
            source_scene_id = safe(ws.cell(row=row, column=source_col).value)
            action = safe(ws.cell(row=row, column=action_col).value)
            outcome = safe(ws.cell(row=row, column=outcome_col).value).upper()
            if not source_scene_id or not action or not outcome:
                continue
            if action == CONTINUE_ACTION and outcome == CONTINUE_OUTCOME:
                continue
            source_scene = scenes.get(source_scene_id)
            if not source_scene or source_scene.scene_type == "static":
                continue
            allowed_pool = split_pool(ws.cell(row=row, column=allowed_col).value)
            if is_transition_already_wrapped(allowed_pool, scenes):
                continue
            transition_id = safe(ws.cell(row=row, column=transition_id_col).value) if transition_id_col else ""
            suffix = stable_suffix(library_key, transition_id, source_scene_id, action, outcome, str(row))
            static_scene_id = f"{STATIC_PREFIX}{library_key}_{suffix}"
            static_transition_id = f"{STATIC_TRANSITION_PREFIX}{library_key}_{suffix}"
            owner_after = safe(ws.cell(row=row, column=owner_col).value) or source_scene.owner
            if owner_after not in SCENE_LIBRARIES:
                continue
            original_next = safe(ws.cell(row=row, column=next_col).value) if next_col else ""
            original_allowed = join_pool(allowed_pool)
            if not original_allowed and original_next:
                original_allowed = original_next
            if not original_allowed:
                continue
            patches.append(
                TransitionPatch(
                    library_key=library_key,
                    source_row=row,
                    source_transition_id=transition_id,
                    source_scene_id=source_scene_id,
                    source_owner=source_scene.owner,
                    action=action,
                    outcome=outcome,
                    ball_owner_after=owner_after,
                    player_position_after=safe(ws.cell(row=row, column=player_pos_after_col).value) if player_pos_after_col else source_scene.player_position,
                    player_direction_after=safe(ws.cell(row=row, column=player_dir_after_col).value) if player_dir_after_col else source_scene.player_direction,
                    ball_holder_position_after=safe(ws.cell(row=row, column=ball_holder_pos_after_col).value) if ball_holder_pos_after_col else source_scene.ball_holder_position,
                    ball_holder_direction_after=safe(ws.cell(row=row, column=ball_holder_dir_after_col).value) if ball_holder_dir_after_col else source_scene.ball_holder_direction,
                    original_next_scene_id=original_next or split_pool(original_allowed)[0],
                    original_allowed_next_scene_ids=original_allowed,
                    static_scene_id=static_scene_id,
                    static_transition_id=static_transition_id,
                    static_text=static_text(source_scene, action, outcome, owner_after),
                )
            )
    return patches


def patch_scene_libraries(patches: List[TransitionPatch]) -> None:
    grouped: Dict[str, List[TransitionPatch]] = {}
    for patch in patches:
        grouped.setdefault(patch.ball_owner_after, []).append(patch)

    for owner, owner_patches in grouped.items():
        path, sheet_name = SCENE_LIBRARIES[owner]
        wb = load_workbook(path)
        ws = wb[sheet_name]
        ensure_column(ws, "scene_type")
        headers = header_map(ws)
        scene_id_col = require(headers, ["scene_id"])
        created = 0
        updated = 0
        for patch in owner_patches:
            existing = find_rows(ws, scene_id_col, patch.static_scene_id)
            if existing:
                row = existing[0]
                updated += 1
            else:
                row = ws.max_row + 1
                created += 1
            values = {
                "scene_id": patch.static_scene_id,
                "scene_owner": patch.ball_owner_after,
                "player_position": patch.player_position_after,
                "player_direction": patch.player_direction_after,
                "ball_holder_position": patch.ball_holder_position_after,
                "ball_holder_direction": patch.ball_holder_direction_after,
                "ball_holder_action": "static event after transition",
                "available_player_actions": "",
                "narrative_scene": patch.static_text,
                "scene_type": "static",
            }
            for col_name, value in values.items():
                set_if_present(ws, headers, row, col_name, value)
        wb.save(path)
        print(f"SCENE_LIBRARY_PATCHED {path.name}: created={created}, updated={updated}")


def patch_transition_libraries(patches: List[TransitionPatch]) -> None:
    grouped: Dict[str, List[TransitionPatch]] = {}
    for patch in patches:
        grouped.setdefault(patch.library_key, []).append(patch)

    for library_key, library_patches in grouped.items():
        path = TRANSITION_LIBRARIES[library_key]
        wb = load_workbook(path)
        ws = wb.active
        headers = header_map(ws)
        transition_id_col = optional(headers, ["transition_id"])
        source_col = require(headers, ["source_scene_id", "source_scene", "from_scene", "scene_id"])
        action_col = require(headers, ["player_action", "action", "teammate_action"])
        outcome_col = require(headers, ["player_action_outcome", "teammate_action_outcome", "outcome"])
        owner_col = require(headers, ["ball_owner_after", "ball_owner"])
        next_col = optional(headers, ["next_scene_id", "target_scene_id"])
        allowed_col = require(headers, ["allowed_next_scene_ids"])
        player_pos_after_col = optional(headers, ["player_position_after"])
        player_dir_after_col = optional(headers, ["player_direction_after"])
        ball_holder_pos_after_col = optional(headers, ["ball_holder_position_after"])
        ball_holder_dir_after_col = optional(headers, ["ball_holder_direction_after"])
        transition_id_to_row: Dict[str, int] = {}
        if transition_id_col is not None:
            for row in range(2, ws.max_row + 1):
                transition_id = safe(ws.cell(row=row, column=transition_id_col).value)
                if transition_id:
                    transition_id_to_row[transition_id] = row
        created = 0
        updated = 0
        redirected = 0
        for patch in library_patches:
            # Re-read row by index; this script does not insert above existing rows.
            source_row = patch.source_row
            ws.cell(row=source_row, column=allowed_col).value = patch.static_scene_id
            if next_col is not None:
                ws.cell(row=source_row, column=next_col).value = patch.static_scene_id
            ws.cell(row=source_row, column=owner_col).value = patch.ball_owner_after
            redirected += 1

            static_row = transition_id_to_row.get(patch.static_transition_id)
            if static_row:
                updated += 1
            else:
                static_row = ws.max_row + 1
                created += 1
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=static_row, column=col).value = ws.cell(row=source_row, column=col).value
                if transition_id_col is not None:
                    transition_id_to_row[patch.static_transition_id] = static_row

            if transition_id_col is not None:
                ws.cell(row=static_row, column=transition_id_col).value = patch.static_transition_id
            ws.cell(row=static_row, column=source_col).value = patch.static_scene_id
            ws.cell(row=static_row, column=action_col).value = CONTINUE_ACTION
            ws.cell(row=static_row, column=outcome_col).value = CONTINUE_OUTCOME
            ws.cell(row=static_row, column=owner_col).value = patch.ball_owner_after
            ws.cell(row=static_row, column=allowed_col).value = patch.original_allowed_next_scene_ids
            if next_col is not None:
                ws.cell(row=static_row, column=next_col).value = patch.original_next_scene_id
            if player_pos_after_col is not None:
                ws.cell(row=static_row, column=player_pos_after_col).value = patch.player_position_after
            if player_dir_after_col is not None:
                ws.cell(row=static_row, column=player_dir_after_col).value = patch.player_direction_after
            if ball_holder_pos_after_col is not None:
                ws.cell(row=static_row, column=ball_holder_pos_after_col).value = patch.ball_holder_position_after
            if ball_holder_dir_after_col is not None:
                ws.cell(row=static_row, column=ball_holder_dir_after_col).value = patch.ball_holder_direction_after
        wb.save(path)
        print(f"TRANSITION_LIBRARY_PATCHED {path.name}: redirected={redirected}, static_created={created}, static_updated={updated}")


def main() -> int:
    scenes = load_scene_records()
    patches = collect_transition_patches(scenes)
    print(f"PEFL static graph patch candidates: {len(patches)}")
    if not patches:
        print("PEFL static graph patch status: PASS (nothing to patch)")
        return 0
    patch_scene_libraries(patches)
    patch_transition_libraries(patches)
    print("PEFL static graph patch status: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
