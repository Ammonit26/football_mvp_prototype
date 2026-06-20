from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

SCENE_LIBRARIES = {
    "PLAYER_WITH_BALL": (ROOT / "ball_at_player_normalized_prototype_v7_executable.xlsx", "normalized_player_scenes"),
    "TEAMMATE_WITH_BALL": (ROOT / "ball_at_teammate_normalized_prototype_v1_executable.xlsx", "normalized_teammate_scenes"),
    "OPPONENT_WITH_BALL": (ROOT / "ball_at_opponent_normalized_prototype_v2_executable.xlsx", "normalized_opponent_scenes"),
}

TRANSITION_LIBRARIES = {
    "player": ROOT / "transitions_player_complete_graph_v2_executable.xlsx",
    "teammate": ROOT / "transitions_teammate_complete_graph_v4_executable.xlsx",
    "opponent": ROOT / "transitions_opponent_complete_graph_v1_executable.xlsx",
}

CONTINUE_ACTION = "CONTINUE"
CONTINUE_OUTCOME = "SUCCESS"


@dataclass(frozen=True)
class StaticPatch:
    source_transition_library: str
    source_scene_id: str
    source_action: str
    source_outcome: str
    static_scene_id: str
    static_transition_id: str
    static_owner: str
    static_text: str
    player_position: str
    player_direction: str
    ball_holder_position: str
    ball_holder_direction: str
    ball_holder_action: str


PATCHES: List[StaticPatch] = [
    StaticPatch(
        source_transition_library="player",
        source_scene_id="fb_player_0002",
        source_action="Пройти соперника дриблингом",
        source_outcome="SUCCESS",
        static_scene_id="fb_static_player_0001",
        static_transition_id="fb_trans_static_player_000001",
        static_owner="PLAYER_WITH_BALL",
        static_text="Ты оставляешь соперника позади и протаскиваешь мяч в центр поля. Атака получает продолжение.",
        player_position="в центре поля",
        player_direction="лицом к чужим воротам",
        ball_holder_position="в центре поля",
        ball_holder_direction="лицом к чужим воротам",
        ball_holder_action="продвигается после успешного дриблинга",
    ),
    StaticPatch(
        source_transition_library="player",
        source_scene_id="fb_player_0002",
        source_action="Пройти соперника дриблингом",
        source_outcome="FAIL",
        static_scene_id="fb_static_player_0002",
        static_transition_id="fb_trans_static_player_000002",
        static_owner="OPPONENT_WITH_BALL",
        static_text="Ты пытаешься пройти соперника у своей штрафной, но он цепляет мяч. Потеря сразу превращается в опасную атаку.",
        player_position="перед своей штрафной",
        player_direction="лицом к чужим воротам",
        ball_holder_position="перед нашей штрафной",
        ball_holder_direction="лицом к чужим воротам",
        ball_holder_action="подхватывает мяч после отбора",
    ),
    StaticPatch(
        source_transition_library="player",
        source_scene_id="fb_player_0158",
        source_action="Отдать пас на фланг",
        source_outcome="SUCCESS",
        static_scene_id="fb_static_player_0003",
        static_transition_id="fb_trans_static_player_000003",
        static_owner="TEAMMATE_WITH_BALL",
        static_text="Пас уходит на фланг точно в темп. Партнёр принимает мяч, отрывается от ближайшего соперника и двигает атаку к штрафной.",
        player_position="на углу чужой штрафной",
        player_direction="лицом к чужим воротам",
        ball_holder_position="перед чужой штрафной",
        ball_holder_direction="лицом к чужим воротам",
        ball_holder_action="принял пас на фланге и продвинулся вперед",
    ),
    StaticPatch(
        source_transition_library="player",
        source_scene_id="fb_player_0158",
        source_action="Отдать пас вперед",
        source_outcome="FAIL",
        static_scene_id="fb_static_player_0004",
        static_transition_id="fb_trans_static_player_000004",
        static_owner="OPPONENT_WITH_BALL",
        static_text="Передача вперед читается защитником. Соперник перехватывает мяч и сразу разворачивает игру в обратную сторону.",
        player_position="в центре поля",
        player_direction="лицом к чужим воротам",
        ball_holder_position="в центре поля",
        ball_holder_direction="лицом к чужим воротам",
        ball_holder_action="перехватил передачу",
    ),
    StaticPatch(
        source_transition_library="teammate",
        source_scene_id="fb_teamm_0065",
        source_action="продвинуться вперед, караулить отскок, рикошет",
        source_outcome="SUCCESS",
        static_scene_id="fb_static_teamm_0001",
        static_transition_id="fb_trans_static_teamm_000001",
        static_owner="TEAMMATE_WITH_BALL",
        static_text="Партнёр находит момент для удара и бьёт низом в дальний угол. Ты читаешь эпизод и двигаешься туда, где может оказаться отскок.",
        player_position="внутри чужой штрафной",
        player_direction="лицом к чужим воротам",
        ball_holder_position="на углу чужой штрафной",
        ball_holder_direction="лицом к чужим воротам",
        ball_holder_action="наносит удар",
    ),
]


def safe(value: object) -> str:
    return "" if value is None else str(value).strip()


def split_pool(raw: object) -> List[str]:
    return [part.strip() for part in safe(raw).split("||") if part.strip()]


def join_pool(values: Iterable[str]) -> str:
    seen: List[str] = []
    for value in values:
        if value and value not in seen:
            seen.append(value)
    return "||".join(seen)


def header_map(ws) -> Dict[str, int]:
    return {safe(cell.value): int(cell.column) for cell in ws[1] if safe(cell.value)}


def require(headers: Dict[str, int], names: Iterable[str]) -> int:
    lowered = {name.lower(): col for name, col in headers.items()}
    for name in names:
        if name.lower() in lowered:
            return lowered[name.lower()]
    raise RuntimeError(f"Missing required column. Tried: {list(names)}")


def optional(headers: Dict[str, int], names: Iterable[str]) -> Optional[int]:
    lowered = {name.lower(): col for name, col in headers.items()}
    for name in names:
        if name.lower() in lowered:
            return lowered[name.lower()]
    return None


def ensure_column(ws, name: str) -> int:
    headers = header_map(ws)
    if name in headers:
        return headers[name]
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = name
    return col


def find_rows(ws, col: int, value: str) -> list[int]:
    return [row for row in range(2, ws.max_row + 1) if safe(ws.cell(row=row, column=col).value) == value]


def set_if_present(ws, headers: Dict[str, int], row: int, column_name: str, value: object) -> None:
    col = optional(headers, [column_name])
    if col is not None:
        ws.cell(row=row, column=col).value = value


def patch_scene_libraries() -> None:
    grouped: Dict[str, List[StaticPatch]] = {}
    for patch in PATCHES:
        grouped.setdefault(patch.static_owner, []).append(patch)

    for owner, patches in grouped.items():
        scene_path, sheet_name = SCENE_LIBRARIES[owner]
        wb = load_workbook(scene_path)
        ws = wb[sheet_name]
        ensure_column(ws, "scene_type")
        headers = header_map(ws)
        scene_id_col = require(headers, ["scene_id"])
        scene_type_col = require(headers, ["scene_type"])

        for row in range(2, ws.max_row + 1):
            if not safe(ws.cell(row=row, column=scene_type_col).value):
                ws.cell(row=row, column=scene_type_col).value = "dynamic"

        for patch in patches:
            existing = find_rows(ws, scene_id_col, patch.static_scene_id)
            if existing:
                row = existing[0]
                status = "updated_existing"
            else:
                row = ws.max_row + 1
                status = "created"

            values = {
                "scene_id": patch.static_scene_id,
                "scene_owner": patch.static_owner,
                "player_position": patch.player_position,
                "player_direction": patch.player_direction,
                "ball_holder_position": patch.ball_holder_position,
                "ball_holder_direction": patch.ball_holder_direction,
                "ball_holder_action": patch.ball_holder_action,
                "available_player_actions": "",
                "narrative_scene": patch.static_text,
                "scene_type": "static",
            }
            for col_name, value in values.items():
                set_if_present(ws, headers, row, col_name, value)
            print(f"SCENE {patch.static_scene_id}: {status} in {scene_path.name}, excel_row={row}")

        wb.save(scene_path)
        print(f"SCENE_LIBRARY_PATCHED {scene_path.name}: {len(patches)} static scene(s)")


def get_static_transition_targets(ws, headers: Dict[str, int], static_transition_id: str) -> List[str]:
    transition_id_col = optional(headers, ["transition_id"])
    allowed_col = require(headers, ["allowed_next_scene_ids"])
    if transition_id_col is None:
        return []
    rows = find_rows(ws, transition_id_col, static_transition_id)
    if not rows:
        return []
    return split_pool(ws.cell(row=rows[0], column=allowed_col).value)


def patch_transition_library(library_key: str, patches: List[StaticPatch]) -> None:
    transition_path = TRANSITION_LIBRARIES[library_key]
    wb = load_workbook(transition_path)
    ws = wb.active
    headers = header_map(ws)
    transition_id_col = optional(headers, ["transition_id"])
    source_col = require(headers, ["source_scene_id", "source_scene", "from_scene", "scene_id"])
    action_col = require(headers, ["player_action", "action", "teammate_action"])
    outcome_col = require(headers, ["player_action_outcome", "teammate_action_outcome", "outcome"])
    owner_col = require(headers, ["ball_owner_after", "ball_owner"])
    next_col = optional(headers, ["next_scene_id", "target_scene_id"])
    allowed_col = require(headers, ["allowed_next_scene_ids"])

    for patch in patches:
        matching_source_rows = []
        template_row = None
        for row in range(2, ws.max_row + 1):
            if (
                safe(ws.cell(row=row, column=source_col).value) == patch.source_scene_id
                and safe(ws.cell(row=row, column=action_col).value) == patch.source_action
                and safe(ws.cell(row=row, column=outcome_col).value).upper() == patch.source_outcome
            ):
                matching_source_rows.append(row)
                if template_row is None:
                    template_row = row

        if not matching_source_rows:
            raise RuntimeError(
                f"No source transition rows found for {patch.source_scene_id} / {patch.source_action} / {patch.source_outcome}"
            )

        current_static_targets = get_static_transition_targets(ws, headers, patch.static_transition_id)
        preserved_targets: List[str] = list(current_static_targets)
        for row in matching_source_rows:
            old_pool = split_pool(ws.cell(row=row, column=allowed_col).value)
            if old_pool == [patch.static_scene_id] and current_static_targets:
                preserved_targets.extend(current_static_targets)
            else:
                preserved_targets.extend(scene_id for scene_id in old_pool if scene_id != patch.static_scene_id)

        if not preserved_targets:
            raise RuntimeError(f"No preserved target pool for static transition {patch.static_transition_id}")

        for row in matching_source_rows:
            ws.cell(row=row, column=allowed_col).value = patch.static_scene_id
            if next_col is not None:
                ws.cell(row=row, column=next_col).value = patch.static_scene_id
            ws.cell(row=row, column=owner_col).value = patch.static_owner

        static_rows = []
        if transition_id_col is not None:
            static_rows = find_rows(ws, transition_id_col, patch.static_transition_id)

        if static_rows:
            static_row = static_rows[0]
            status = "updated_existing"
        else:
            static_row = ws.max_row + 1
            status = "created"
            for col in range(1, ws.max_column + 1):
                ws.cell(row=static_row, column=col).value = ws.cell(row=template_row, column=col).value

        if transition_id_col is not None:
            ws.cell(row=static_row, column=transition_id_col).value = patch.static_transition_id
        ws.cell(row=static_row, column=source_col).value = patch.static_scene_id
        ws.cell(row=static_row, column=action_col).value = CONTINUE_ACTION
        ws.cell(row=static_row, column=outcome_col).value = CONTINUE_OUTCOME
        ws.cell(row=static_row, column=owner_col).value = patch.static_owner
        ws.cell(row=static_row, column=allowed_col).value = join_pool(preserved_targets)
        if next_col is not None:
            ws.cell(row=static_row, column=next_col).value = preserved_targets[0]

        print(
            f"TRANSITION {patch.static_transition_id}: {status}; "
            f"{len(matching_source_rows)} source row(s) redirected to {patch.static_scene_id}; "
            f"continue_pool={join_pool(preserved_targets)}"
        )

    wb.save(transition_path)
    print(f"TRANSITION_LIBRARY_PATCHED {transition_path.name}: {len(patches)} static slice(s)")


def patch_transition_libraries() -> None:
    grouped: Dict[str, List[StaticPatch]] = {}
    for patch in PATCHES:
        grouped.setdefault(patch.source_transition_library, []).append(patch)

    for library_key, patches in grouped.items():
        patch_transition_library(library_key, patches)


def main() -> int:
    patch_scene_libraries()
    patch_transition_libraries()
    print("Static scene package patch status: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
