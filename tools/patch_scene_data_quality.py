from __future__ import annotations

from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook

PATCHES: Dict[Tuple[str, str, str], str] = {
    (
        "ball_at_teammate_normalized_prototype_v1_executable.xlsx",
        "fb_teamm_0081",
        "ball_holder_action",
    ): "Осматривается в поиске партнеров",
    (
        "ball_at_opponent_normalized_prototype_v2_executable.xlsx",
        "fb_opp_0061",
        "ball_holder_action",
    ): "Держит мяч под давлением",
    (
        "ball_at_opponent_normalized_prototype_v2_executable.xlsx",
        "fb_opp_0061",
        "narrative_scene",
    ): "Соперник находится в нашей штрафной. Соперник: держит мяч под давлением. Персонаж находится перед своей штрафной лицом к чужим воротам.",
    (
        "ball_at_opponent_normalized_prototype_v2_executable.xlsx",
        "fb_opp_0063",
        "ball_holder_action",
    ): "Держит мяч под давлением",
    (
        "ball_at_opponent_normalized_prototype_v2_executable.xlsx",
        "fb_opp_0063",
        "narrative_scene",
    ): "Соперник находится в нашей штрафной. Соперник: держит мяч под давлением. Персонаж находится в центре поля лицом к чужим воротам.",
}


def find_header_map(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = int(cell.column)
    return headers


def find_scene_row(ws, scene_id_col: int, scene_id: str) -> int:
    matches = []
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=scene_id_col).value
        if str(value).strip() == scene_id:
            matches.append(row_idx)
    if len(matches) != 1:
        raise RuntimeError(f"Expected exactly one row for {scene_id}, found {len(matches)}")
    return matches[0]


def row_snapshot(ws, row_idx: int) -> Dict[int, object]:
    return {col_idx: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)}


def apply_file_patches(file_name: str, file_patches: Dict[Tuple[str, str], str]) -> None:
    path = Path(file_name)
    if not path.exists():
        raise FileNotFoundError(path)

    wb = load_workbook(path)
    ws = wb.active
    headers = find_header_map(ws)
    if "scene_id" not in headers:
        raise RuntimeError(f"Missing required column 'scene_id' in {file_name}")

    scene_id_col = headers["scene_id"]
    touched = []

    for (scene_id, column_name), new_value in file_patches.items():
        if column_name not in headers:
            raise RuntimeError(f"Missing required column {column_name!r} in {file_name}")

        target_col = headers[column_name]
        row_idx = find_scene_row(ws, scene_id_col, scene_id)
        before = row_snapshot(ws, row_idx)
        old_value = ws.cell(row=row_idx, column=target_col).value
        ws.cell(row=row_idx, column=target_col).value = new_value
        after = row_snapshot(ws, row_idx)

        changed_columns = [col for col in before if before[col] != after[col]]
        if old_value == new_value:
            changed_columns = []
        elif changed_columns != [target_col]:
            raise RuntimeError(
                f"Unexpected changes while patching {scene_id}: changed columns {changed_columns}, expected {[target_col]}"
            )

        touched.append((scene_id, column_name, old_value, new_value, row_idx, bool(changed_columns)))

    wb.save(path)
    print(f"PATCHED FILE {file_name}")
    for scene_id, column_name, old_value, new_value, row_idx, changed in touched:
        status = "changed" if changed else "already_ok"
        print(f"  {scene_id} :: {column_name} :: {status}")
        print(f"    old: {old_value}")
        print(f"    new: {new_value}")
        print(f"    excel_row: {row_idx}")
    print("  verification: only target cells changed before save")


def main() -> int:
    patches_by_file: Dict[str, Dict[Tuple[str, str], str]] = {}
    for (file_name, scene_id, column_name), new_value in PATCHES.items():
        patches_by_file.setdefault(file_name, {})[(scene_id, column_name)] = new_value

    for file_name, file_patches in patches_by_file.items():
        apply_file_patches(file_name, file_patches)

    print("Scene data quality patch status: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
